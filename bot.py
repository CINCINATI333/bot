import vk_api
import random
import time
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import json

COLUMN_WIDTH = 4
TOKEN = "токен"
itheration_num = 1

def get_rand():
  return random.random()*1000000

def send( id, text ):
    global itheration_num

    log = open( "lor.txt", 'a' )
    #log.write( json.dumps( {"peer_id": id, "message": text, "random_id": get_rand()} ) )
    payload = {"peer_id": id, "message": text, "random_id": get_rand()}
    payback = vk.method( "messages.send", payload ) 

    log.write( 
        json.dumps( 
            {
                "payload": payload,
                "payback": payback,
                "time": time.ctime( time.time() ),
                "itheration_num": itheration_num
            }             
        ) + "\n"
    )
    log.close()
    time.sleep( 1 )
    return

def itheration():
    global itheration_num

    messages = vk.method("messages.getConversations", {"offset": 0, "count": 20, "filter": "unread"})   #ответ вк
    if messages and messages["count"]:
        conversations = messages["items"]
        for conv in conversations:
            print(conv)
            id = conv["conversation"]["peer"]["id"] #Получили ID приславшего сообщение 
            text = conv["last_message"]["text"] #Записали  переменную текст сообщения   
            in_read_id = conv["conversation"]["out_read"] #id последнего прочитанного входящего
            
            uid = id
            fields = "first_name, last_name"
            res = vk.method("users.get", {"user_ids": uid, "fields": fields })
            user = { 'name': 'xname', 'surname': 'xsurname' }
            uname = res[0]['first_name'] #Записали в переменную имя пользователя
            ulastname = res[0]['last_name'] #Записали в переенную фамилию пользователя

            if text.lower() == "омае ва му":
                send( id, "шиндейру!" )
            else:
                send( id, "Nein!!!!" )  
            text = text.split() #Разбили текст сообщения на массив
            print(text)
            try:
                match = str(text[0])
                score1 = str(text[1])
                score2 = str(text[2])
            except Exception:
                print( "Wrong line syntax!" )
                send( id, "Какая-то ошибка, попробуйте ещё раз или обратитесь в службу поддержки пользователей!" )
                return

            unumber = 4 #Разница в количестве столбцов между ячейками с номерами матчей
            wbook = xlrd.open_workbook('2.xls') #Открыл файл Excel
            wb = copy(wbook)
            w_sheet = wb.get_sheet(0) #открыл первый лист
            w_sheet2 = wbook.sheet_by_index(0) #Открыл первый лист другим методом (другой библиотекой)
            b = float(w_sheet2.cell(0,0).value)
            print(b)

            match=match+'.0'
            a = [0 for i in range(1,100)]
            k=0
            i=1
            while i<30:
                a[i] = str(w_sheet2.cell(0,int(i)).value) #записть в переменную данных с 1-ой строки экселя (строка, столбец)
                if a[i] == match: #если значение в 1 строке столбца i равно переменной match (присланноый номер мачта)
                    k=i
                    unumber = float(b)+4 #номер строки, в первом столбце которой будет имя и фамилия пользователя
                    w_sheet.write(0,0,float(b)+1) #Запись в ячейку количества пользователей, отправивших сообщения
                    b+=1
                i+=1
            
            w_sheet.write(unumber,k+2,text[1]+ '-' + text[2]) #запись счёта
            w_sheet.write(unumber,0,uname +' '+ ulastname) #запись имени и фамилии
            w_sheet.write(0,0 ,b) #запись количества пользователей, приславших сообщение
            wb.save('2.xls')

            
            print(res[0]['first_name']) #Имя
            print(res[0]['last_name']) #Фамилия

            print(text[0],'  - первый элемент сообщения')
            print(text[1], ' - второй элемент сообщения')
            print(text[2], ' - третий элемент сообщения')

    itheration_num += 1
    print( str(itheration_num) + " итерация" )
    time.sleep( 4 )
    return



vk = vk_api.VkApi( token = TOKEN )
while True:
    itheration()
      

   
